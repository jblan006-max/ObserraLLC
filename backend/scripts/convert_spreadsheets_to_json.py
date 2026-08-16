"""Convert downloaded NIST supplemental spreadsheets into a basic JSON controls list.

This script expects files downloaded into deploy/onprem/nist_docs and will attempt to parse known
workbooks to extract control identifiers and titles. Output file is deploy/onprem/sp800_53_rev5_extracted.json
which can be imported with backend/scripts/import_nist.py --file <path> or via the import-local endpoint.

Note: Spreadsheet formats vary; this script performs a best-effort extraction and should be validated.
"""
import os
import json
from openpyxl import load_workbook

# compute repository root (one level above the 'backend' package)
repo_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
in_dir = os.path.join(repo_root, "deploy", "onprem", "nist_docs")
out_file = os.path.join(repo_root, "deploy", "onprem", "sp800_53_rev5_extracted.json")

controls = []
if not os.path.exists(in_dir):
    print("No nist_docs directory found at", in_dir)
    exit(1)

for f in os.listdir(in_dir):
    if not f.lower().endswith(('.xlsx', '.xlsm', '.xls')):
        continue
    path = os.path.join(in_dir, f)
    print("Processing", path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=1, max_row=5000, values_only=True):
                if not row:
                    continue
                # heuristics: rows with control ids like 'AC-1' or 'IA-2(1)'
                for cell in row:
                    if isinstance(cell, str) and '-' in cell and any(ch.isdigit() for ch in cell):
                        txt = cell.strip()
                        # crude match
                        if len(txt) <= 20 and any(part.isalpha() for part in txt.split('-')):
                            controls.append({"id": txt, "title": ""})
        wb.close()
    except Exception as e:
        print("Failed processing", path, e)

# dedupe
unique = {}
for c in controls:
    cid = c.get('id')
    if cid and cid not in unique:
        unique[cid] = c

out = list(unique.values())
with open(out_file, 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)

print(f"Wrote {len(out)} extracted controls to {out_file}")
